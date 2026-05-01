import streamlit as st
import pandas as pd
import requests
import time
import io
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Facilio Readings Onboarder",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

.stApp {
    background: #0a0e1a;
    color: #e2e8f0;
}

h1, h2, h3 {
    font-family: 'IBM Plex Mono', monospace !important;
    letter-spacing: -0.5px;
}

.header-bar {
    background: linear-gradient(135deg, #0f1729 0%, #1a2744 100%);
    border: 1px solid #2d4a8a;
    border-radius: 12px;
    padding: 28px 36px;
    margin-bottom: 32px;
    position: relative;
    overflow: hidden;
}
.header-bar::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #3b82f6, #06b6d4, #8b5cf6);
}
.header-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.8rem;
    font-weight: 600;
    color: #f1f5f9;
    margin: 0 0 6px 0;
}
.header-sub {
    color: #64748b;
    font-size: 0.9rem;
    margin: 0;
}

.phase-card {
    background: #0f1729;
    border: 1px solid #1e3a5f;
    border-radius: 10px;
    padding: 24px;
    margin-bottom: 20px;
}
.phase-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    font-weight: 600;
    color: #3b82f6;
    text-transform: uppercase;
    letter-spacing: 2px;
    margin-bottom: 16px;
}

.hint-box {
    background: #0d1f3c;
    border: 1px solid #1e3a5f;
    border-left: 3px solid #3b82f6;
    border-radius: 6px;
    padding: 12px 16px;
    font-size: 0.82rem;
    color: #94a3b8;
    margin-top: 8px;
    font-family: 'IBM Plex Mono', monospace;
}

.stat-row {
    display: flex;
    gap: 12px;
    margin: 16px 0;
}
.stat-box {
    flex: 1;
    background: #0f1729;
    border-radius: 8px;
    padding: 14px 18px;
    text-align: center;
}
.stat-num {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem;
    font-weight: 600;
}
.stat-label {
    font-size: 0.75rem;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 1px;
}
.green { color: #22c55e; border: 1px solid #14532d; }
.red   { color: #ef4444; border: 1px solid #7f1d1d; }
.yellow{ color: #f59e0b; border: 1px solid #78350f; }
.blue  { color: #3b82f6; border: 1px solid #1e3a5f; }
.grey  { color: #94a3b8; border: 1px solid #1e293b; }

.warn-box {
    background: #1c1008;
    border: 1px solid #92400e;
    border-radius: 8px;
    padding: 14px 18px;
    color: #fbbf24;
    font-size: 0.88rem;
    margin: 12px 0;
}
.error-box {
    background: #1c0a0a;
    border: 1px solid #7f1d1d;
    border-radius: 8px;
    padding: 14px 18px;
    color: #f87171;
    font-size: 0.88rem;
    margin: 12px 0;
}
.success-box {
    background: #0a1c0f;
    border: 1px solid #14532d;
    border-radius: 8px;
    padding: 14px 18px;
    color: #4ade80;
    font-size: 0.88rem;
    margin: 12px 0;
}
.info-box {
    background: #0d1f3c;
    border: 1px solid #1e3a5f;
    border-radius: 8px;
    padding: 14px 18px;
    color: #93c5fd;
    font-size: 0.88rem;
    margin: 12px 0;
}

.batch-card {
    background: #0f1729;
    border: 1px solid #1e3a5f;
    border-radius: 10px;
    padding: 20px 24px;
    margin: 10px 0;
}

/* Streamlit overrides */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea {
    background: #0f1729 !important;
    border: 1px solid #1e3a5f !important;
    color: #e2e8f0 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.85rem !important;
    border-radius: 6px !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 2px rgba(59,130,246,0.2) !important;
}
label, .stCheckbox label {
    color: #94a3b8 !important;
    font-size: 0.85rem !important;
}
.stButton > button {
    background: #1d4ed8 !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important;
    letter-spacing: 0.5px !important;
    padding: 10px 24px !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: #2563eb !important;
    transform: translateY(-1px) !important;
}
.stDownloadButton > button {
    background: #065f46 !important;
    color: #ecfdf5 !important;
    border: 1px solid #10b981 !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important;
}
div[data-testid="stNumberInput"] input {
    background: #0f1729 !important;
    border: 1px solid #1e3a5f !important;
    color: #e2e8f0 !important;
}
.stProgress > div > div {
    background: #3b82f6 !important;
}
div[data-testid="stExpander"] {
    background: #0f1729;
    border: 1px solid #1e3a5f !important;
    border-radius: 8px !important;
}
</style>
""", unsafe_allow_html=True)


# ─── Session state init ───────────────────────────────────────────────────────
def init_state():
    defaults = {
        "connected": False,
        "org_name": "",
        "org_id": "",
        "df": None,
        "df_sorted": None,
        "categories_map": {},       # displayName.lower() → {id, assetModuleID}
        "metric_map": {},           # unit_symbol.lower() → {metricId, unitId, label}
        "validated": False,
        "validation_errors": [],
        "validation_warnings": [],
        "unmatched_categories": [],
        "processed_rows": [],       # rows ready to send with resolved IDs
        "pilot_done": False,
        "pilot_results": [],
        "batch_index": 0,           # which batch we're on (0-based, after pilot)
        "all_results": [],
        "run_complete": False,
        "paused": False,
        "session": None,
        "base_url": "",
        "headers": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ─── Constants ────────────────────────────────────────────────────────────────
READING_TYPE_MAP = {
    "decimal": 3,
    "number":  1,
    "boolean": 2,
    "string":  4,
    "enum":    5,
}

# Unit symbol → (metric label hint for matching)
# We build the real map after fetching /getDefaultMetricUnits
UNIT_HINT_MAP = {
    # Electrical
    "a": "current", "amp": "current", "ampere": "current",
    "v": "voltage", "volt": "voltage",
    "w": "power", "watt": "power",
    "kw": "power",
    "kwh": "energy", "kilowatthour": "energy", "kwh": "energy",
    "mwh": "energy",
    "var": "reactive power", "kvar": "reactive power",
    "va": "apparent power", "kva": "apparent power",
    "hz": "frequency",
    "mω": "resistance", "mohm": "resistance",
    "ω": "resistance", "ohm": "resistance",
    "f": "capacitance",
    # Temperature
    "°c": "temperature", "c": "temperature", "celsius": "temperature",
    "°f": "temperature", "f": "temperature",
    "k": "temperature",
    # Pressure
    "bar": "pressure", "pa": "pressure", "kpa": "pressure",
    "psi": "pressure", "mbar": "pressure",
    # Flow
    "m³/h": "flow", "m3/h": "flow", "lph": "flow", "lpm": "flow",
    "m³/s": "flow", "l/s": "flow",
    # Humidity
    "%rh": "humidity", "rh": "humidity",
    # Speed
    "rpm": "angular velocity", "m/s": "velocity",
    # Volume / Level
    "l": "volume", "m³": "volume", "m3": "volume",
    "%": "dimensionless",
    # Time
    "h": "time", "s": "time", "min": "time",
    # Misc
    "db": "sound level", "ph": "ph", "μs/cm": "conductivity",
    "mg/l": "concentration", "mm": "length", "m": "length",
}


# ─── API helpers ──────────────────────────────────────────────────────────────
def build_headers(cookie: str, csrf: str, org_id: str) -> dict:
    return {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Cookie": cookie,
        "X-Csrf-Token": csrf,
        "X-Org-Id": str(org_id),
        "X-Org-Group": "v2",
        "X-Version": "revive",
        "X-Device-Type": "Web",
        "X-App-Version": "v1816",
        "X-current-site": "-1",
    }


def fetch_account(base_url: str, headers: dict):
    r = requests.get(
        f"{base_url}/maintenance/api/v2/fetchAccount",
        headers=headers, timeout=15
    )
    return r.json()


def fetch_all_categories(base_url: str, headers: dict) -> dict:
    """Paginate through ALL asset categories dynamically."""
    categories = {}
    page = 1
    per_page = 50
    total = None

    while True:
        r = requests.get(
            f"{base_url}/maintenance/api/v3/modules/assetcategory",
            params={"page": page, "perPage": per_page, "withCount": "true", "moduleName": "assetcategory"},
            headers=headers, timeout=15
        )
        data = r.json()
        if data.get("code", -1) != 0:
            break
        items = data["data"].get("assetcategory", [])
        if total is None:
            total = data["meta"]["pagination"]["totalCount"]
        for cat in items:
            key = (cat.get("displayName") or "").strip().lower()
            categories[key] = {
                "id": cat["id"],
                "assetModuleID": cat.get("assetModuleID"),
                "displayName": cat.get("displayName", ""),
                "name": cat.get("name", ""),
            }
        fetched_so_far = (page - 1) * per_page + len(items)
        if fetched_so_far >= total or not items:
            break
        page += 1

    return categories, total


def fetch_metric_units(base_url: str, headers: dict) -> dict:
    """Returns symbol.lower() → {metricId, unitId, displayName, metricName}"""
    r = requests.get(
        f"{base_url}/maintenance/api/setup/units/getDefaultMetricUnits",
        headers=headers, timeout=15
    )
    data = r.json()
    unit_map = {}
    metrics_meta = data.get("metrics", {})
    for metric_name, units_list in data.get("metricWithUnits", {}).items():
        metric_info = metrics_meta.get(metric_name, {})
        metric_id = metric_info.get("metricId", -1)
        for u in units_list:
            symbol = (u.get("symbol") or "").strip().lower()
            display = (u.get("displayName") or "").strip().lower()
            unit_id = u.get("unitId", -1)
            entry = {
                "metricId": metric_id,
                "unitId": unit_id,
                "displayName": u.get("displayName", ""),
                "symbol": u.get("symbol", ""),
                "metricName": metric_name,
                "siUnit": u.get("siUnit", False),
            }
            if symbol:
                unit_map[symbol] = entry
            if display and display != symbol:
                unit_map[display] = entry
    return unit_map


def get_existing_readings(base_url: str, headers: dict, category_id: int) -> set:
    """Returns set of existing reading displayNames (lowercased) for a category."""
    r = requests.get(
        f"{base_url}/maintenance/api/v2/readings/assetcategory",
        params={"id": category_id, "excludeEmptyFields": "false",
                "readingType": "available", "fetchValidationRules": "true"},
        headers=headers, timeout=15
    )
    data = r.json()
    readings = data.get("result", {}).get("readings", [])
    return {(rd.get("displayName") or "").strip().lower() for rd in readings}


def resolve_metric_unit(unit_str: str, metric_map: dict) -> dict:
    """Given a unit string from Excel, find best matching metricId + unitId."""
    if not unit_str or str(unit_str).strip() in ("", "nan"):
        return {"metricId": -1, "unitId": -1, "symbol": ""}
    key = str(unit_str).strip().lower()
    if key in metric_map:
        return metric_map[key]
    # Try partial match
    for k, v in metric_map.items():
        if key in k or k in key:
            return v
    return {"metricId": -1, "unitId": -1, "symbol": unit_str}


def post_reading(base_url: str, headers: dict, category_id: int,
                 reading_name: str, data_type: int,
                 metric_id: int, unit_id: int) -> dict:
    field_json = {
        "displayName": reading_name,
        "dataType": data_type,
        "dataTypeTemp": data_type,
        "counterField": False,
        "safeLimitPattern": "none",
        "raiseSafeLimitAlarm": False,
        "safeLimitSeverity": "Minor",
        "inputPatternSeverity": "Minor",
        "lesserThan": None,
        "greaterThan": None,
        "betweenTo": None,
        "betweenFrom": None,
        "livePointStatus": False,
        "livePointDisplayName": "",
        "livePointAggregation": None,
        "livePointInterval": None,
    }
    if metric_id and metric_id != -1:
        field_json["metric"] = metric_id
    if unit_id and unit_id != -1:
        field_json["unit"] = unit_id

    payload = {
        "resourceType": "Asset",
        "parentCategoryId": category_id,
        "readingName": reading_name,
        "fieldJsons": [field_json],
        "sensorRuleList": [{
            "sensorRuleTypes": [],
            "sensorAlarmDetails": {"message": "", "severity": ""},
            "selectedValues": [],
        }],
        "fieldReadingRules": [[]],
    }
    r = requests.post(
        f"{base_url}/maintenance/api/setup/reading/addsetupreading",
        json=payload, headers=headers, timeout=15
    )
    return r.json()


# ─── Excel export ─────────────────────────────────────────────────────────────
def build_result_excel(all_results: list, pilot_results: list) -> bytes:
    combined = pilot_results + all_results
    wb = Workbook()

    fills = {
        "SUCCESS":   PatternFill("solid", start_color="0A2E17"),
        "FAILED":    PatternFill("solid", start_color="2E0A0A"),
        "DUPLICATE": PatternFill("solid", start_color="2E2000"),
        "SKIPPED":   PatternFill("solid", start_color="1A1A2E"),
    }
    fonts = {
        "SUCCESS":   Font(color="4ADE80", name="Calibri"),
        "FAILED":    Font(color="F87171", name="Calibri"),
        "DUPLICATE": Font(color="FBBf24", name="Calibri"),
        "SKIPPED":   Font(color="94A3B8", name="Calibri"),
    }
    header_fill = PatternFill("solid", start_color="0F1729")
    header_font = Font(color="3B82F6", bold=True, name="Calibri")

    def write_sheet(ws, rows, columns):
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(c, "") for c in columns])
            status = row.get("Status", "")
            if status in fills:
                for cell in ws[ws.max_row]:
                    cell.fill = fills[status]
                    cell.font = fonts[status]
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    counts = {"SUCCESS": 0, "FAILED": 0, "DUPLICATE": 0, "SKIPPED": 0}
    for r in combined:
        counts[r.get("Status", "SKIPPED")] = counts.get(r.get("Status", "SKIPPED"), 0) + 1
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["Total Processed", len(combined)])
    ws_summary.append(["✅ Success",   counts["SUCCESS"]])
    ws_summary.append(["❌ Failed",    counts["FAILED"]])
    ws_summary.append(["⏭ Duplicate", counts["DUPLICATE"]])
    ws_summary.append(["⏸ Skipped",   counts["SKIPPED"]])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 12

    cols = ["Row", "Asset Category", "Reading Display Name", "Reading Type", "Unit", "Status", "Error"]

    for title, status_filter in [("Success","SUCCESS"),("Failed","FAILED"),("Duplicate","DUPLICATE"),("Skipped","SKIPPED")]:
        rows = [r for r in combined if r.get("Status") == status_filter]
        if rows:
            ws = wb.create_sheet(title)
            write_sheet(ws, rows, cols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Validation ───────────────────────────────────────────────────────────────
def validate_dataframe(df: pd.DataFrame):
    errors = []
    warnings = []
    required = ["Asset Category", "Reading Display Name", "Reading Type"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        errors.append(f"Missing required columns: {', '.join(missing_cols)}")
        return errors, warnings

    blank_names = df[df["Reading Display Name"].isna() | (df["Reading Display Name"].astype(str).str.strip() == "")].index.tolist()
    if blank_names:
        errors.append(f"Blank 'Reading Display Name' in {len(blank_names)} rows: {blank_names[:5]}{'...' if len(blank_names)>5 else ''}")

    blank_cats = df[df["Asset Category"].isna() | (df["Asset Category"].astype(str).str.strip() == "")].index.tolist()
    if blank_cats:
        errors.append(f"Blank 'Asset Category' in {len(blank_cats)} rows: {blank_cats[:5]}{'...' if len(blank_cats)>5 else ''}")

    valid_types = set(READING_TYPE_MAP.keys())
    invalid_types = df[~df["Reading Type"].astype(str).str.strip().str.lower().isin(valid_types)]["Reading Type"].dropna().unique().tolist()
    if invalid_types:
        errors.append(f"Invalid Reading Type values: {invalid_types}. Allowed: {list(valid_types)}")

    if "Unit" in df.columns:
        blank_units = df["Unit"].isna().sum()
        if blank_units > 0:
            warnings.append(f"{blank_units} rows have no Unit — readings will be created without metric/unit assignment.")

    return errors, warnings


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="header-bar">
  <div class="header-title">⚡ Facilio Readings Mass Onboarder</div>
  <p class="header-sub">Bulk-configure new readings into Facilio — validated, batched, and confirmed at every step.</p>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 0 — CONNECTION
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="phase-card">', unsafe_allow_html=True)
st.markdown('<div class="phase-title">Phase 0 — Connection Setup</div>', unsafe_allow_html=True)

col1, col2 = st.columns([1, 1])
with col1:
    base_url = st.text_input("Facilio Base URL", placeholder="https://app.facilio.co.ae",
                              value=st.session_state.get("base_url", ""))
    cookie = st.text_area("Cookie String", placeholder="fc.session=...; fc.csrfToken=...",
                           height=80, help="Copy from DevTools → Network → any /api/ request → Request Headers → Cookie")
with col2:
    csrf = st.text_input("X-Csrf-Token", placeholder="63-character hex string",
                          type="password",
                          help="Same network request → X-Csrf-Token header value")
    st.markdown("""<div class="hint-box">
    📌 <b>How to get Cookie & CSRF Token:</b><br>
    1. Log into Facilio in Chrome<br>
    2. Press <b>F12</b> → Network tab<br>
    3. Reload the page — click any <code>/api/</code> request<br>
    4. Headers tab → scroll to <b>Request Headers</b><br>
    5. Copy <code>Cookie</code> and <code>X-Csrf-Token</code> values
    </div>""", unsafe_allow_html=True)

if st.button("🔌 Test Connection", disabled=not (base_url and cookie and csrf)):
    with st.spinner("Connecting to Facilio..."):
        try:
            headers_temp = build_headers(cookie, csrf, "0")
            result = fetch_account(base_url.rstrip("/"), headers_temp)
            if result.get("responseCode") == 0:
                org = result["result"]["account"]["org"]
                org_id = str(org.get("id", ""))
                org_name = org.get("name", "Unknown")
                st.session_state.connected = True
                st.session_state.org_id = org_id
                st.session_state.org_name = org_name
                st.session_state.base_url = base_url.rstrip("/")
                st.session_state.headers = build_headers(cookie, csrf, org_id)
                st.rerun()
            else:
                st.markdown(f'<div class="error-box">❌ Auth failed — responseCode: {result.get("responseCode")}. Check your Cookie and CSRF token.</div>', unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f'<div class="error-box">❌ Connection error: {e}</div>', unsafe_allow_html=True)

if st.session_state.connected:
    st.markdown(f'<div class="success-box">✅ Connected to <b>{st.session_state.org_name}</b> &nbsp;|&nbsp; Org ID: <code>{st.session_state.org_id}</code></div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — UPLOAD & VALIDATE (only if connected)
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.connected:
    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 1 — Upload & Validate Excel</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader("Upload Readings Excel", type=["xlsx", "xls"],
                                 help="Must contain: Asset Category, Reading Display Name, Reading Type, Unit columns")

    if uploaded:
        df_raw = pd.read_excel(uploaded)
        # Sort by Asset Category
        if "Asset Category" in df_raw.columns:
            df_sorted = df_raw.sort_values("Asset Category", na_position="last").reset_index(drop=True)
        else:
            df_sorted = df_raw.reset_index(drop=True)

        st.session_state.df = df_raw
        st.session_state.df_sorted = df_sorted

        errors, warnings = validate_dataframe(df_sorted)
        st.session_state.validation_errors = errors
        st.session_state.validation_warnings = warnings

        col_a, col_b = st.columns([1, 1])
        with col_a:
            st.markdown(f"**📄 File:** `{uploaded.name}`")
            st.markdown(f"**Rows:** `{len(df_sorted)}` &nbsp;|&nbsp; **Columns:** `{list(df_sorted.columns)}`")
        with col_b:
            if errors:
                for e in errors:
                    st.markdown(f'<div class="error-box">❌ {e}</div>', unsafe_allow_html=True)
            if warnings:
                for w in warnings:
                    st.markdown(f'<div class="warn-box">⚠️ {w}</div>', unsafe_allow_html=True)
            if not errors:
                st.markdown('<div class="success-box">✅ Validation passed — file is ready.</div>', unsafe_allow_html=True)
                st.session_state.validated = True

        if not errors:
            with st.expander("👁 Preview sorted data (first 10 rows)"):
                st.dataframe(df_sorted.head(10), use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — FETCH FACILIO DATA & CROSS-CHECK
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.connected and st.session_state.validated:
    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 2 — Fetch Facilio Data & Cross-Check</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"**Categories fetched:** {'`' + str(len(st.session_state.categories_map)) + '`' if st.session_state.categories_map else '_not fetched yet_'}")
    with col2:
        st.markdown(f"**Metric units loaded:** {'`' + str(len(st.session_state.metric_map)) + '`' if st.session_state.metric_map else '_not fetched yet_'}")

    if st.button("🔄 Fetch Categories & Metric Units from Facilio"):
        with st.spinner("Fetching all asset categories (paginated)..."):
            try:
                cats, total = fetch_all_categories(
                    st.session_state.base_url, st.session_state.headers
                )
                st.session_state.categories_map = cats
                st.markdown(f'<div class="success-box">✅ Fetched <b>{len(cats)}</b> categories (total in Facilio: {total})</div>', unsafe_allow_html=True)
            except Exception as e:
                st.markdown(f'<div class="error-box">❌ Failed to fetch categories: {e}</div>', unsafe_allow_html=True)

        with st.spinner("Fetching metric/unit reference data..."):
            try:
                mmap = fetch_metric_units(st.session_state.base_url, st.session_state.headers)
                st.session_state.metric_map = mmap
                st.markdown(f'<div class="success-box">✅ Loaded <b>{len(mmap)}</b> unit/metric mappings</div>', unsafe_allow_html=True)
            except Exception as e:
                st.markdown(f'<div class="warn-box">⚠️ Could not fetch metric units: {e} — readings will be created without metric assignment.</div>', unsafe_allow_html=True)
        st.rerun()

    # Cross-check once we have both
    if st.session_state.categories_map and st.session_state.df_sorted is not None:
        df = st.session_state.df_sorted
        cats = st.session_state.categories_map
        mmap = st.session_state.metric_map

        unmatched = []
        processed = []

        for idx, row in df.iterrows():
            cat_name = str(row.get("Asset Category", "")).strip()
            reading_name = str(row.get("Reading Display Name", "")).strip()
            reading_type = str(row.get("Reading Type", "Decimal")).strip().lower()
            unit_str = str(row.get("Unit", "")).strip() if pd.notna(row.get("Unit")) else ""

            cat_key = cat_name.lower()
            cat_info = cats.get(cat_key)

            if not cat_info:
                unmatched.append({"row": idx + 2, "category": cat_name})
                continue

            data_type = READING_TYPE_MAP.get(reading_type, 3)
            metric_info = resolve_metric_unit(unit_str, mmap)

            processed.append({
                "row_index": idx,
                "row_num": idx + 2,
                "Asset Category": cat_name,
                "Reading Display Name": reading_name,
                "Reading Type": reading_type,
                "Unit": unit_str,
                "category_id": cat_info["id"],
                "data_type": data_type,
                "metricId": metric_info.get("metricId", -1),
                "unitId": metric_info.get("unitId", -1),
                "metric_symbol": metric_info.get("symbol", ""),
            })

        st.session_state.processed_rows = processed
        st.session_state.unmatched_categories = unmatched

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(f'<div class="success-box">✅ <b>{len(processed)}</b> rows matched and ready to process.</div>', unsafe_allow_html=True)
        with col_b:
            if unmatched:
                st.markdown(f'<div class="warn-box">⚠️ <b>{len(unmatched)}</b> rows have unrecognised Asset Categories — will be skipped.</div>', unsafe_allow_html=True)
                with st.expander(f"View {len(unmatched)} unmatched category rows"):
                    st.dataframe(pd.DataFrame(unmatched), use_container_width=True)
            else:
                st.markdown('<div class="success-box">✅ All categories matched in Facilio.</div>', unsafe_allow_html=True)

        # Show resolved metric preview
        if processed:
            with st.expander("🔍 Preview resolved metric/unit mapping (first 15 rows)"):
                preview_df = pd.DataFrame(processed[:15])[
                    ["row_num", "Asset Category", "Reading Display Name", "Unit", "metric_symbol", "metricId", "unitId"]
                ]
                st.dataframe(preview_df, use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.connected and st.session_state.validated and st.session_state.processed_rows:
    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 3 — Run Configuration</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        delay_ms = st.number_input("Delay between API calls (ms)", min_value=100, max_value=2000,
                                    value=300, step=50,
                                    help="Prevents rate-limiting. 300ms recommended.")
    with col2:
        skip_duplicates = st.checkbox("Skip duplicate readings", value=True,
                                       help="Check if reading already exists before POSTing. Recommended ON.")
    with col3:
        batch_size = st.number_input("Batch size", min_value=10, max_value=100, value=50, step=10,
                                      help="Rows per batch. You'll confirm each batch before it runs.")

    st.session_state["delay_sec"] = delay_ms / 1000
    st.session_state["skip_duplicates"] = skip_duplicates
    st.session_state["batch_size"] = int(batch_size)

    total_ready = len(st.session_state.processed_rows)
    unmatched_count = len(st.session_state.unmatched_categories)
    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-box blue"><div class="stat-num">{total_ready}</div><div class="stat-label">Ready to Process</div></div>
      <div class="stat-box grey"><div class="stat-num">{unmatched_count}</div><div class="stat-label">Will Be Skipped</div></div>
      <div class="stat-box blue"><div class="stat-num">{(total_ready - 3 + int(batch_size) - 1) // int(batch_size)}</div><div class="stat-label">Batches (after pilot)</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 4 — PILOT RUN (3 readings)
# ══════════════════════════════════════════════════════════════════════════════
def run_rows(rows_to_run, skip_dup, delay_sec):
    """Execute API calls for a list of processed rows. Returns list of result dicts."""
    results = []
    readings_cache = {}  # category_id → set of existing reading names

    for row in rows_to_run:
        cat_id = row["category_id"]
        reading_name = row["Reading Display Name"]

        # Duplicate check
        if skip_dup:
            if cat_id not in readings_cache:
                try:
                    readings_cache[cat_id] = get_existing_readings(
                        st.session_state.base_url, st.session_state.headers, cat_id
                    )
                except:
                    readings_cache[cat_id] = set()

            if reading_name.strip().lower() in readings_cache[cat_id]:
                results.append({**row, "Status": "DUPLICATE", "Error": "Already exists in Facilio"})
                time.sleep(delay_sec)
                continue

        # POST
        try:
            resp = post_reading(
                st.session_state.base_url, st.session_state.headers,
                row["category_id"], reading_name,
                row["data_type"], row["metricId"], row["unitId"]
            )
            if resp.get("responseCode") == 0 or resp.get("code") == 0:
                results.append({**row, "Status": "SUCCESS", "Error": ""})
                # Update cache
                if cat_id in readings_cache:
                    readings_cache[cat_id].add(reading_name.strip().lower())
            else:
                err = resp.get("message") or resp.get("errorMessage") or str(resp)
                results.append({**row, "Status": "FAILED", "Error": err})
        except Exception as e:
            results.append({**row, "Status": "FAILED", "Error": str(e)})

        time.sleep(delay_sec)

    return results


if (st.session_state.connected and st.session_state.validated
        and st.session_state.processed_rows and not st.session_state.pilot_done):

    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 4 — Pilot Run (First 3 Readings)</div>', unsafe_allow_html=True)

    pilot_rows = st.session_state.processed_rows[:3]
    st.markdown("**The following 3 readings will be created as a pilot:**")
    pilot_preview = pd.DataFrame(pilot_rows)[["row_num", "Asset Category", "Reading Display Name", "Unit", "metric_symbol"]]
    st.dataframe(pilot_preview, use_container_width=True)

    if st.button("🚀 Run Pilot (3 readings)"):
        delay = st.session_state.get("delay_sec", 0.3)
        skip = st.session_state.get("skip_duplicates", True)
        with st.spinner("Running pilot..."):
            pilot_results = run_rows(pilot_rows, skip, delay)
        st.session_state.pilot_results = pilot_results

        s = sum(1 for r in pilot_results if r["Status"] == "SUCCESS")
        f = sum(1 for r in pilot_results if r["Status"] == "FAILED")
        d = sum(1 for r in pilot_results if r["Status"] == "DUPLICATE")

        st.markdown(f"""
        <div class="stat-row">
          <div class="stat-box green"><div class="stat-num">{s}</div><div class="stat-label">Success</div></div>
          <div class="stat-box red"><div class="stat-num">{f}</div><div class="stat-label">Failed</div></div>
          <div class="stat-box yellow"><div class="stat-num">{d}</div><div class="stat-label">Duplicate</div></div>
        </div>
        """, unsafe_allow_html=True)

        if f > 0:
            for r in pilot_results:
                if r["Status"] == "FAILED":
                    st.markdown(f'<div class="error-box">❌ Row {r["row_num"]}: {r["Reading Display Name"]} — {r["Error"]}</div>', unsafe_allow_html=True)

        st.markdown("""
        <div class="warn-box">
        ⚠️ <b>Please verify in Facilio</b> that the pilot readings were added correctly before continuing.<br>
        Go to: Setup → Portfolio Settings → Readings → find the categories above.
        </div>
        """, unsafe_allow_html=True)

        st.rerun()

    if st.session_state.pilot_results:
        s = sum(1 for r in st.session_state.pilot_results if r["Status"] == "SUCCESS")
        f = sum(1 for r in st.session_state.pilot_results if r["Status"] == "FAILED")
        d = sum(1 for r in st.session_state.pilot_results if r["Status"] == "DUPLICATE")

        st.markdown(f"**Pilot result:** ✅ {s} success &nbsp; ❌ {f} failed &nbsp; ⏭ {d} duplicate")
        st.markdown("**Did the pilot additions look correct in Facilio?**")
        col_yes, col_no = st.columns([1, 4])
        with col_yes:
            if st.button("✅ Yes — proceed to batches"):
                st.session_state.pilot_done = True
                st.session_state.batch_index = 0
                st.rerun()
        with col_no:
            if st.button("❌ No — stop and download report"):
                skipped = [
                    {**r, "Status": "SKIPPED", "Error": "Stopped after failed pilot"}
                    for r in st.session_state.processed_rows[3:]
                ]
                skipped += [
                    {**r, "Status": "SKIPPED", "Error": "Unmatched category"}
                    for r in st.session_state.unmatched_categories
                ]
                excel_bytes = build_result_excel(skipped, st.session_state.pilot_results)
                st.download_button("⬇️ Download Partial Report", data=excel_bytes,
                                    file_name="facilio_onboard_partial.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 5 — BATCH PROCESSING
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.pilot_done and not st.session_state.run_complete:
    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 5 — Batch Processing</div>', unsafe_allow_html=True)

    remaining = st.session_state.processed_rows[3:]  # exclude pilot rows
    batch_size = st.session_state.get("batch_size", 50)
    batches = [remaining[i:i + batch_size] for i in range(0, len(remaining), batch_size)]
    total_batches = len(batches)
    batch_idx = st.session_state.batch_index

    # Overall progress
    total_done = len(st.session_state.all_results)
    total_remaining = len(remaining)
    if total_remaining > 0:
        progress = total_done / total_remaining
        st.progress(progress)
        st.markdown(f"**Overall:** `{total_done}` / `{total_remaining}` rows processed (excluding pilot)")

    if batch_idx < total_batches and not st.session_state.paused:
        current_batch = batches[batch_idx]
        st.markdown(f"""
        <div class="batch-card">
        <b>Batch {batch_idx + 1} of {total_batches}</b> &nbsp;|&nbsp;
        Rows: <code>{current_batch[0]['row_num']}</code> → <code>{current_batch[-1]['row_num']}</code> &nbsp;|&nbsp;
        Size: <code>{len(current_batch)}</code>
        </div>
        """, unsafe_allow_html=True)

        if st.button(f"▶ Run Batch {batch_idx + 1} of {total_batches}"):
            delay = st.session_state.get("delay_sec", 0.3)
            skip = st.session_state.get("skip_duplicates", True)

            progress_bar = st.progress(0)
            status_text = st.empty()

            batch_results = []
            for i, row in enumerate(current_batch):
                status_text.markdown(f"Processing row `{row['row_num']}` — **{row['Reading Display Name']}**")
                result = run_rows([row], skip, delay)
                batch_results.extend(result)
                progress_bar.progress((i + 1) / len(current_batch))

            st.session_state.all_results.extend(batch_results)

            s = sum(1 for r in batch_results if r["Status"] == "SUCCESS")
            f = sum(1 for r in batch_results if r["Status"] == "FAILED")
            d = sum(1 for r in batch_results if r["Status"] == "DUPLICATE")

            st.markdown(f"""
            <div class="stat-row">
              <div class="stat-box green"><div class="stat-num">{s}</div><div class="stat-label">Success</div></div>
              <div class="stat-box red"><div class="stat-num">{f}</div><div class="stat-label">Failed</div></div>
              <div class="stat-box yellow"><div class="stat-num">{d}</div><div class="stat-label">Duplicate</div></div>
            </div>
            """, unsafe_allow_html=True)

            if f > 0:
                with st.expander(f"❌ {f} failed rows in this batch"):
                    for r in batch_results:
                        if r["Status"] == "FAILED":
                            st.markdown(f"- Row `{r['row_num']}` **{r['Reading Display Name']}**: {r['Error']}")

            st.session_state.batch_index += 1
            if st.session_state.batch_index >= total_batches:
                st.session_state.run_complete = True
            st.rerun()

        col_pause, col_stop = st.columns([1, 4])
        with col_pause:
            if st.button("⏸ Pause"):
                st.session_state.paused = True
                st.rerun()

    elif st.session_state.paused:
        st.markdown('<div class="warn-box">⏸ Run is paused. Resume or stop below.</div>', unsafe_allow_html=True)
        col_res, col_stop = st.columns([1, 4])
        with col_res:
            if st.button("▶ Resume"):
                st.session_state.paused = False
                st.rerun()
        with col_stop:
            if st.button("⏹ Stop & Download Report"):
                st.session_state.run_complete = True
                st.rerun()

    elif batch_idx >= total_batches:
        st.session_state.run_complete = True
        st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 6 — FINAL REPORT
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.run_complete:
    st.markdown('<div class="phase-card">', unsafe_allow_html=True)
    st.markdown('<div class="phase-title">Phase 6 — Final Report</div>', unsafe_allow_html=True)

    all_combined = st.session_state.pilot_results + st.session_state.all_results
    skipped_unmatched = [
        {"row_num": r["row"], "Asset Category": r["category"],
         "Reading Display Name": "", "Reading Type": "", "Unit": "",
         "Status": "SKIPPED", "Error": "Asset category not found in Facilio"}
        for r in st.session_state.unmatched_categories
    ]

    s = sum(1 for r in all_combined if r["Status"] == "SUCCESS")
    f = sum(1 for r in all_combined if r["Status"] == "FAILED")
    d = sum(1 for r in all_combined if r["Status"] == "DUPLICATE")
    sk = len(skipped_unmatched)

    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-box green"><div class="stat-num">{s}</div><div class="stat-label">✅ Success</div></div>
      <div class="stat-box red"><div class="stat-num">{f}</div><div class="stat-label">❌ Failed</div></div>
      <div class="stat-box yellow"><div class="stat-num">{d}</div><div class="stat-label">⏭ Duplicate</div></div>
      <div class="stat-box grey"><div class="stat-num">{sk}</div><div class="stat-label">⏸ Skipped</div></div>
    </div>
    """, unsafe_allow_html=True)

    excel_bytes = build_result_excel(all_combined + skipped_unmatched, [])
    st.download_button(
        label="⬇️ Download Full Result Report (.xlsx)",
        data=excel_bytes,
        file_name="facilio_readings_onboard_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if f > 0:
        with st.expander(f"❌ View {f} failed readings"):
            failed = [r for r in all_combined if r["Status"] == "FAILED"]
            st.dataframe(pd.DataFrame(failed)[["row_num", "Asset Category", "Reading Display Name", "Error"]], use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)
